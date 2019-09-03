
from django.shortcuts import render ,HttpResponse
import os
import pandas as pd
import openpyxl
import csv
from utility.FileOperating import FileOperating
import sys
import numpy as np

input_dir = r"D:\merge_data\input_data"
output_dir = r"D:\merge_data\output_data"

def csv_merge_searching(request):
    return render(request, 'assist/csv_merge.html')

def write_excel_xlsx(path,sheet_name,value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0,index):
        for j in range(0,len(value[i])):
            try:
                sheet.cell(row = i + 1,column = j +1 , value = float(value[i][j]))
            except:
                sheet.cell(row = i + 1,column = j +1 , value = str(value[i][j]))
    workbook.save(path)

def add_sheet_xlsx(output_final, sheet_name,value,title2):
    path = str(output_final + '\\' + title2 + '.xlsx')
    wb = openpyxl.load_workbook(path)
    ws2 = wb.create_sheet(sheet_name)
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            try:
                ws2.cell(row=i + 1, column=j + 1, value=float(value[i][j]))
            except:
                ws2.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        wb.save(path)

# def merge_data(name1,input_final,output_final,title):
#     output_path = output_final + '\\' + title+ '.xlsx'
#     for k in range(len(name1)):
#         input_path = open(input_final + '\\' + name1[k],'rt',encoding='gb18030',errors='ignore')
#         df = csv.reader(input_path)
#         print(df)
#         li1 = []
#         for i in df:
#             li1.append(i)
#         if k==0:
#             write_excel_xlsx(output_path,name1[k],li1)
#         else:
#             add_sheet_xlsx(output_final, name1[k], li1, title)

def merge_data(name1,input_final,output_final,title):
    output_path = output_final + '\\' + title+ '.xlsx'
    for k in range(len(name1)):
        input_path = input_final + '\\' + name1[k]
        df = pd.read_excel(input_path)
        df1 = np.array(df)
        df2 = df1.tolist()
        if k==0:
            write_excel_xlsx(output_path,name1[k],df2)
        else:
            add_sheet_xlsx(output_final, name1[k], df2, title)

def csv_merge_load(request):
    user_id = request.session.get('user_id')
    if request.method =='POST':
        # 创建原始数据文件路径
        f = FileOperating(input_dir, user_id)
        f.root_creating()
        input_final, input_date, input_time = f.file_creating()

        f = FileOperating(output_dir, user_id)
        f.root_creating()
        output_final, output_date, output_time = f.file_creating()

        myFiles = request.FILES.getlist('myfile', None)
        if not myFiles:
            return render(request, 'assist/csv_merge.html', {'upload_result': '请选择文件'})

        for myFile in myFiles:
            destination = open(os.path.join(input_final, myFile.name), 'wb+')
            for chunk in myFile.chunks():
                destination.write(chunk)
            destination.close()
        write = pd.ExcelWriter(output_final + '\\' + 'summary.xls')

        names = os.listdir(input_final)
        name1 = []
        for name in names:
            name1.append(name)
        title = '汇总文件'       # 文件名

        merge_data(name1, input_final,output_final, title)
        results = {'date': output_date,'time':output_time, 'user':user_id,'name':str(title) + '.xlsx'}
        return render(request,'assist/csv_merge_resluts.html',{'results':results})

def csv_merge_date_searching(request):
    user_id = request.session.get('user_id')
    f = FileOperating(output_dir +'\\'+ user_id,user_id)
    result = f.file_searching()
    return render(request,'assist/merge_searching_result.html',{'result':result})

def csv_merge_time_searching(request):
    date = request.GET.get('date')
    user_id = request.session.get('user_id')
    f = FileOperating(output_dir +'\\'+ user_id+ '\\'+ date,user_id)
    result = f.file_searching()
    return render(request, 'assist/merge_searching_result1.html', {'result': result,'date':date})
