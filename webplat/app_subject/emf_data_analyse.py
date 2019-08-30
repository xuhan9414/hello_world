from django.shortcuts import render
from django.shortcuts import HttpResponse
import os
import pandas as pd
import  matplotlib.pyplot as plt
from utility.FileOperating import FileOperating
import numpy as np
import openpyxl
import copy
import  datetime
import matplotlib.ticker as ticker

input_dir = r"D:\emf_testing_data\input_data"
output_dir = r"D:\emf_testing_data\output_data"

def write_excel_xlsx(path,sheet_name,value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0,index):
        for j in range(0,len(value[i])):
            try:
                sheet.cell(row = i + 1,column = j +1 , value = float(value[i][j]))
            except:
                sheet.cell(row = i + 1,column = j +1 , value = str(value[i][j]))
    workbook.save(path)

def add_sheet_xlsx(output_final, sheet_name,value,title2):
    path = str(output_final + '\\' + title2 + '.xlsx')
    wb = openpyxl.load_workbook(path)
    ws2 = wb.create_sheet(sheet_name)
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            try:
                ws2.cell(row=i + 1, column=j + 1, value=float(value[i][j]))
            except:
                ws2.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        wb.save(path)

def string_toDatetime(st):
    a = datetime.datetime.strftime(st,"%m-%d %H:%M:%S")
    return a

def input_data(name1,input_final,output_final,m,e1,e2,t,z,title2):
    for i in name1:
        input_path = open(input_final + '\\' + i ,'rt')
        df = pd.read_csv(input_final,skiprows=[0,1,2,3],encoding='UTF-8')
        labels = list(df.columns.value)
        df_li = df.values.tolist()
        df_times = []
        for item in df_li:
            df_times.append(item[0])
            del item[0]

        t2 = len(df_li[0])
        if t2<t:
            return 0

        for i in range(len(df_li)):
            for j in range(t):
                df_li[i][j] = 10 ** (df_li[i][j] / 100)

        df_sum = copy.deepcopy(df_li)
        df_sum = np.array(df_sum)
        df_sum = np.transpose(df_sum)
        df_sum1 = []
        for i in range(len(df_sum[0,:])):
            dg1 = np.sum(df_sum[:,i])
            dg1 = 100 * np.log10(dg1)/10
            df_sum1.append(dg1)

        df_datetime = []
        for i in df_datetime:
            df_datetime.append(string_toDatetime(i))
        delt_time = (df_datetime[1]-df_datetime[0]).seconds

        df_li = np.array(df_li)
        li2 = [[]for i in range(t)]
        for i in range(t):
            for j in range(len(df_li)-int(e1/delt_time - 1)):
                li2[i].append(np.mean(df_li[j:j+int(e1/delt_time),i]))

        li2 = np.array(li2)
        testsum1 = []
        for i in range(len(li2[0,:])):
            testsum1.append(np.nansum(li2[:,i]))
        testsum2 = []
        for i in testsum1:
            if i == 0 :
                testsum2.append(0)
            else:
                testsum2.append(100*np.log10(i)/10)
        radio1 = []
        for i in testsum2:
            radio1.append(i - z/10)

        df_out1 = copy.deepcopy(df_li)
        df_out1 = np.array(df_out1)
        df_out1 = df_out1.tolist()
        for i in range(len(df_out1)):
            df_out1[i].insert(0, df_times[i])
        df_out1.insert(0, labels)

        label1 = copy.deepcopy(labels)
        df_out2 = copy.deepcopy(li2)
        df_out2 = np.array(df_out2)
        df_out2 = np.transpose(df_out2)
        df_out2 = df_out2.tolist()
        for i in range(len(df_out2)):
            df_out2[i].insert(0, df_times[i+int(e1/delt_time-1)])
        df_out2.insert(0,label1)
        df_out2[0][0] = "每通道60s平均"

        li3 = [list(i) for i in zip(testsum1,testsum2,radio1)]
        label2 = ['多通道60s平均','多通道平均(mW)','多通道总和(0.1dBm)','折算时隙配比']
        df_out3 = copy.deepcopy(li3)
        for i in range(len(df_out3)):
            df_out3[i].insert(0,df_times[i+int(e1/delt_time-1)])
        df_out3.insert(0,label2)

        output_path = output_final + '\\' + title2 +'.xlsx'
        write_excel_xlsx(output_path , 'data_sum' , df_out3)
        add_sheet_xlsx(output_final,'power(mW)',df_out1 , title2)
        add_sheet_xlsx(output_final,'data(mW)' , df_out2,title2)

        plt.rcParams['font.sans-serif'] = ['SimHei']
        plt.rcParams['axes.unicode_minus'] = False

        df_times1 = []
        for i in df_times:
            b = list(i.split())
            df_times1.append(b[1])
        x = df_times1[int(e1/delt_time-1):]
        x1 = df_times1[:]
        tick_spacing = len(x1)/5

        m = 10 ** (m / 100)*t
        # e2 = 10**(e2/100)
        # tgp = m - e2
        tgp = 100 * np.log10(m)/10
        tgp = tgp -e2/10
        tgp1 = []
        for i in range(len(x1)):
            tgp1.append(tgp)

        fig,ax = plt.subplots(1,1)
        plt.plot(x1,df_sum1,color = 'blue',label = 'Cell Power every second')
        plt.plot(x1,tgp1,'--',color = 'coral',label = 'Target Power')
        plt.plot(x, testsum2, 'kp-',color = 'magenta',ms=0.05,label='Cell Average Power in EMF Period')
        plt.plot(x, radio1,'o',color = 'green',ms=1,label = '折算时隙配比')
        plt.ylabel('Cell Average output Pwr(dBm)')
        plt.title(title2)
        plt.legend(loc='lower left' , framealpha = 0)
        ax.xaxis.set_major_locator(ticker.MultipleLocator(tick_spacing))
        plt.ylim(tgp-30,tgp+20)
        plt.xticks(rotation = 30)

        max_indx = np.argmax(radio1)
        plt.plot(x[max_indx],radio1[max_indx])
        show_max = ' '+ 'max: ' +str(round(radio1[max_indx,2]))+'(targetpower:'+str(round(tgp,2))+')'
        if radio1[max_indx]>tgp:
            plt.annotate(show_max,xytext = (max_indx+15,radio1[max_indx]-20),xy=(x[max_indx],radio1[max_indx]),
                         arrowprops = dict(arrowstyle='->'))

        plt.savefig(output_final+'/result.png')
        return 1

def analyse_emf(request):
    user = request.session.get('user_id')

    myFiles = request.FILES.getlist('myfile', None)
    if not myFiles:
        return render(request,'subject/emf_analyse.html',{'upload_result':'请选择文件'})
    maxtrp = request.POST.get('maxtrp')
    emfperiod = request.POST.get('emfperiod')
    maxthrd = request.POST.get('maxthrd')
    trxport = request.POST.get('trxport')
    if not maxtrp:
        return render(request,'subject/emf_analyse.html',{'upload_result':'请输入maxtrp'})
    if not emfperiod:
        return render(request,'subject/emf_analyse.html',{'upload_result':'请输入emfperiod'})
    if not maxthrd:
        return render(request,'subject/emf_analyse.html',{'upload_result':'请输入maxthrd'})
    if not trxport:
        return render(request,'subject/emf_analyse.html',{'upload_result':'请输入trxport'})
    else:
        maxtrp = int(maxtrp)
        emfperiod = int(emfperiod)
        maxthrd = int(maxthrd)
        trxport = int(trxport)
    zzpb = request.POST.get('zzpb')
    if zzpb == '1':
        zpb = 12.9
    if zzpb == '2':
        zpb = 12.9

    f = FileOperating(input_dir,user)
    f.root_creating()
    input_final,input_date,input_time = f.file_creating()

    f = FileOperating(output_dir, user)
    f.root_creating()
    output_final,output_date, output_time = f.file_creating()

    fn = request.FILES.get('myfile')
    fn = str(fn)
    global title2
    if "_柜号" in fn:
        title1 = fn.split('_柜号')
        title2 = title1[0]
    else:
        title2 = 'NRCell Average Power'

    for myFile in myFiles:
        destination = open(os.path.join(input_final , myFile.name),'wb+')
        for chunk in myFile.chunks():
            destination.write(chunk)
        destination.close()

    names = os.listdir(input_final)
    name1 = []
    for name in names:
        name1.append(name)

    a = input_data(name1,input_final,output_final,maxtrp,emfperiod,maxthrd,trxport,zpb,title2)
    if a ==0 :
        return render(request,'subject/emf_analyse.html',{"upload_result":'TRXPort 大于文件通道数'})
    results = {"rank_mcs":"result.png","date":output_date,'time':output_time,'user':user,
               'name':(str(title2)+'.xlsx')}
    return render(request, 'subject/emf_resluts.html', {"results": results})
